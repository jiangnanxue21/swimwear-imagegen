"""字段映射、草稿校验与导出(阶段 2 的 2D)。

阶段 2 退出指标里有三条落在这个文件上:

    必填字段漏检   0 次   -> 缺必填必须阻断导出
    导出一致率     100%   -> Excel/CSV/预览三者逐格相同
    属性不得静默流出       -> 只有 CONFIRMED 的属性能进平台字段
"""
from __future__ import annotations

from app.attributes.contracts import AttrValue, CanonicalProduct, CanonicalSku
from app.channels import generic
from app.channels.spec_loader import ChannelSpecError, load_spec_dict
from app.core.enums import (
    AttributeSource,
    AttributeStatus,
    DraftStatus,
    ImageSetStatus,
    MediaRole,
)
from app.listings import export_writer
from app.listings.contracts import (
    CopySnapshot,
    ImageItemSnapshot,
    ImageSetSnapshot,
    ListingDraftData,
)
from app.listings.draft import can_export, copy_locale_problems
from tests.pure._helpers import BACKEND_ROOT  # noqa: F401  (装 sys.path)


def _attr(name: str, value, *, status=AttributeStatus.CONFIRMED) -> AttrValue:
    return AttrValue(
        field_name=name,
        value=value,
        source=AttributeSource.MANUAL,
        status=status,
        version_id=f"v-{name}",
    )


#: 这份夹具的两行 SKU 同属一个颜色(NVY),图片在 SPU 通用位上。
#:
#: A45-batch24 起行级图片只从 `color_sku_image_map` 翻译,所以这里给一份
#: 显式映射:两行都指向那张主图。本文件验的是表头映射、变换、校验与导出
#: 写盘,行级图片的挑选口径由 `test_export_variant_and_field_types.py`
#: 与 batch24 的接缝守卫管。
_MAP = {
    "schema": "1",
    "colors": {
        "NVY": {
            "variant_code": "NVY",
            "skus": {
                "SW-001-NVY-S": {"primary": "m1", "extras": ["m2"]},
                "SW-001-NVY-M": {"primary": "m1", "extras": ["m2"]},
            },
        }
    },
}


def _draft(**overrides) -> ListingDraftData:
    attrs = {
        "primary_color": _attr("primary_color", "NAVY_BLUE"),
        "garment_type": _attr("garment_type", "ONE_PIECE"),
        "material": _attr("material", "82% Nylon 18% Spandex"),
        "neckline_type": _attr("neckline_type", "V_NECK"),
    }
    attrs.update(overrides.pop("attrs", {}))

    product = CanonicalProduct(
        spu="SW-001",
        category_path=("Swimwear",),
        spu_attrs=attrs,
        skus=(
            CanonicalSku(sku="SW-001-NVY-S", variant_id="NVY", size="S"),
            CanonicalSku(sku="SW-001-NVY-M", variant_id="NVY", size="M"),
        ),
    )
    image_set = ImageSetSnapshot(
        image_set_id="is-1",
        spu="SW-001",
        version=2,
        status=ImageSetStatus.APPROVED,
        items=(
            ImageItemSnapshot(
                media_asset_id="m1",
                role=MediaRole.PRODUCT_FRONT,
                sort_order=0,
                storage_path="media/front.jpg",
                width=1200,
                height=1600,
                is_primary=True,
            ),
            ImageItemSnapshot(
                media_asset_id="m2",
                role=MediaRole.PRODUCT_BACK,
                sort_order=1,
                storage_path="media/back.jpg",
                width=1200,
                height=1600,
            ),
        ),
    )
    copies = {
        "en": CopySnapshot(
            copy_id="c1",
            locale="en",
            version=3,
            title="Navy Blue One Piece Swimsuit",
            bullet_points=("Soft stretch fabric", "V neckline", "Full back"),
            description="A navy blue one piece swimsuit made from nylon blend.",
            keywords=("swimsuit", "one piece"),
        )
    }
    data = dict(
        spu="SW-001",
        channel=generic.CHANNEL,
        site=generic.SITE,
        category_id=generic.CATEGORY_ID,
        product=product,
        image_set=image_set,
        copies=copies,
        manual={
            "brand": "Marine",
            "price": "39.90",
            "currency": "USD",
            "stock": "120",
            "lead_time_days": "5",
        },
        spec_version="1.0.0",
        mapping_version="1",
    )
    data.update(overrides)
    return ListingDraftData(**data)


def _spec():
    return generic.field_spec()


# ---------------------------------------------------------------- spec 加载


def test_generic_spec_loads_and_has_no_todo_fields():
    """SHEIN 那份整份是 TODO 加载即拒;这一份必须能真的加载。"""
    spec = _spec()
    assert spec.channel == "GENERIC"
    assert spec.site == "MAIN"
    assert spec.todos() == ()
    assert spec.header_fields and spec.row_fields


def test_spec_rejects_a_variant_reference_in_a_spu_field():
    """SPU 级字段引用 variant.* 是配置错误,不是运行期取不到值。

    后者会安静地产出一个空单元格,而空单元格要到平台驳回才被发现。
    """
    raw = {
        "channel": "X",
        "sites": ["MAIN"],
        "header_fields": [{"key": "a", "source": "variant.sku"}],
    }
    try:
        load_spec_dict(raw, site="MAIN", category_id="swimwear")
    except ChannelSpecError as exc:
        assert "SKU 行" in str(exc)
    else:
        raise AssertionError("SPU 级字段引用 variant.* 应当被拒绝")


def test_spec_rejects_an_unparseable_source_expression():
    raw = {
        "channel": "X",
        "sites": ["MAIN"],
        "header_fields": [{"key": "a", "source": "copy.en.title.upper()"}],
    }
    try:
        load_spec_dict(raw, site="MAIN", category_id="swimwear")
    except ChannelSpecError:
        # 吞的是**期望中的那一个**:下面的 `else` 在它没抛时失败。
        # `tests/pure/` 不许 import pytest,所以没有 `pytest.raises`。
        pass
    else:
        raise AssertionError("非法 source 表达式应当在加载期被拒绝")


def test_spec_rejects_duplicate_field_keys():
    raw = {
        "channel": "X",
        "sites": ["MAIN"],
        "header_fields": [
            {"key": "a", "source": "product.spu"},
            {"key": "a", "source": "product.spu"},
        ],
    }
    try:
        load_spec_dict(raw, site="MAIN", category_id="swimwear")
    except ChannelSpecError as exc:
        assert "重复" in str(exc)
    else:
        raise AssertionError("重复字段 key 应当被拒绝")


# ---------------------------------------------------------------- 映射


def test_mapping_fills_header_and_one_row_per_sku():
    mapped = generic.map_fields(_draft(), _spec(), image_map=_MAP)
    assert mapped.header["spu_code"] == "SW-001"
    assert mapped.header["product_title"] == "Navy Blue One Piece Swimsuit"
    assert mapped.header["garment_type"] == "ONE_PIECE"
    assert mapped.header["main_image"] == "media/front.jpg"
    assert mapped.header["back_image"] == "media/back.jpg"
    assert len(mapped.rows) == 2
    assert [r["size"] for r in mapped.rows] == ["S", "M"]
    assert all(r["price"] == "39.90" for r in mapped.rows)


def test_bullet_points_are_joined_by_the_declared_separator():
    mapped = generic.map_fields(_draft(), _spec(), image_map=_MAP)
    assert mapped.header["selling_points"] == (
        "Soft stretch fabric | V neckline | Full back"
    )


def test_unconfirmed_attributes_never_reach_platform_fields():
    """§8.1:只有 CONFIRMED 的属性能进事实源。

    这一条不能靠调用方记得过滤 —— 它是映射层的不变量。
    """
    draft = _draft(
        attrs={
            "material": _attr(
                "material", "guessed fabric", status=AttributeStatus.SUGGESTED
            )
        }
    )
    mapped = generic.map_fields(draft, _spec(), image_map=_MAP)
    assert mapped.header["material"] is None

    problems = generic.validate(mapped, _spec())
    assert any(
        p.field_key == "material" and p.code == "CHANNEL_FIELD_MISSING"
        for p in problems
    )


def test_title_is_truncated_to_the_declared_limit():
    long_title = "N" * 300
    draft = _draft(
        copies={
            "en": CopySnapshot(
                copy_id="c1", locale="en", version=1, title=long_title,
                description="d" * 20,
            )
        }
    )
    mapped = generic.map_fields(draft, _spec(), image_map=_MAP)
    assert len(mapped.header["product_title"]) == 120


def test_primary_image_follows_the_editorial_sort_order():
    """同一角色多张图时取排序最靠前的 —— 排序是运营的编辑决定。"""
    draft = _draft()
    extra = ImageItemSnapshot(
        media_asset_id="m9",
        role=MediaRole.PRODUCT_FRONT,
        sort_order=5,
        storage_path="media/front-alt.jpg",
        width=800,
        height=1000,
    )
    swapped = ImageSetSnapshot(
        image_set_id=draft.image_set.image_set_id,
        spu=draft.image_set.spu,
        version=draft.image_set.version,
        status=draft.image_set.status,
        items=(extra, *draft.image_set.items),
    )
    mapped = generic.map_fields(_draft(image_set=swapped), _spec(), image_map=_MAP)
    assert mapped.header["main_image"] == "media/front.jpg"


def test_disabled_images_are_not_exported():
    draft = _draft()
    disabled = ImageSetSnapshot(
        image_set_id="is-1",
        spu="SW-001",
        version=1,
        status=ImageSetStatus.APPROVED,
        items=tuple(
            ImageItemSnapshot(
                media_asset_id=i.media_asset_id,
                role=i.role,
                sort_order=i.sort_order,
                storage_path=i.storage_path,
                width=i.width,
                height=i.height,
                enabled=False,
            )
            for i in draft.image_set.items
        ),
    )
    mapped = generic.map_fields(_draft(image_set=disabled), _spec(), image_map=_MAP)
    assert mapped.header["main_image"] is None


# ---------------------------------------------------------------- 校验


def test_a_clean_draft_has_no_blocking_violations():
    mapped = generic.map_fields(_draft(), _spec(), image_map=_MAP)
    problems = generic.validate(mapped, _spec())
    blocking = [p for p in problems if p.is_blocking]
    assert blocking == [], [p.message for p in blocking]


def test_missing_manual_price_blocks_export():
    """验收脚本第 10 步:制造必填字段缺失 -> 草稿不能正式导出。"""
    draft = _draft(manual={"brand": "Marine", "currency": "USD", "stock": "1"})
    mapped = generic.map_fields(draft, _spec(), image_map=_MAP)
    problems = generic.validate(mapped, _spec())
    missing = [p for p in problems if p.code == "CHANNEL_FIELD_MISSING"]
    assert {p.field_key for p in missing} == {"price"}
    # 两行 SKU 都要各报一次,否则「哪一行缺」无法回答
    assert sorted(p.row_index for p in missing) == [0, 1]
    assert not MappedIsValid(mapped, problems)


def MappedIsValid(mapped, problems) -> bool:
    return not any(p.is_blocking for p in problems)


def test_enum_value_outside_the_allowed_list_is_rejected():
    draft = _draft(attrs={"garment_type": _attr("garment_type", "SPACESUIT")})
    mapped = generic.map_fields(draft, _spec(), image_map=_MAP)
    problems = generic.validate(mapped, _spec())
    assert any(p.code == "CHANNEL_FIELD_NOT_ALLOWED" for p in problems)


def test_a_draft_without_any_sku_row_is_rejected():
    product = CanonicalProduct(spu="SW-001", spu_attrs={}, skus=())
    mapped = generic.map_fields(_draft(product=product), _spec(), image_map=_MAP)
    problems = generic.validate(mapped, _spec())
    assert any(p.code == "CHANNEL_NO_ROWS" for p in problems)


def test_image_count_check_is_separate_from_the_spec():
    mapped = generic.map_fields(_draft(), _spec(), image_map=_MAP)
    assert generic.image_count_check(mapped, minimum=2) == []
    assert generic.image_count_check(mapped, minimum=4)


# ---------------------------------------------------------------- 导出一致性


def _meta():
    return export_writer.ExportMeta(
        spu="SW-001",
        channel=generic.CHANNEL,
        site=generic.SITE,
        category_id=generic.CATEGORY_ID,
        spec_version="1.0.0",
    )


def test_csv_cells_match_the_preview_cell_for_cell():
    """阶段 2 退出指标:导出一致率 100%。

    这条测试是那个指标的机器版本。它比「人工比对一遍」强的地方在于
    每次改映射都会重跑。
    """
    spec = _spec()
    mapped = generic.map_fields(_draft(), spec, image_map=_MAP)
    preview = export_writer.preview_tables(mapped, spec)

    import csv as _csv
    import io as _io

    rows = list(_csv.reader(_io.StringIO(export_writer.to_csv(mapped, spec))))
    header_row, *data_rows = rows

    spu_labels = [f["label"] for f in preview["header"]]
    sku_labels = [f["label"] for f in preview["rows"][0]["fields"]]
    assert header_row == [*spu_labels, *sku_labels]

    spu_values = [f["value"] or "" for f in preview["header"]]
    for index, line in enumerate(data_rows):
        assert line[: len(spu_values)] == spu_values
        sku_values = [f["value"] or "" for f in preview["rows"][index]["fields"]]
        assert line[len(spu_values):] == sku_values


def test_json_export_matches_the_preview_values():
    spec = _spec()
    mapped = generic.map_fields(_draft(), spec, image_map=_MAP)
    preview = export_writer.preview_tables(mapped, spec)
    payload = export_writer.to_json(mapped, spec, _meta())

    for field in preview["header"]:
        expected = field["value"]
        actual = payload["header"][field["key"]]
        assert (actual if actual is None else str(actual)) == expected


def test_preview_marks_the_offending_field_not_the_whole_draft():
    """FE-233 的验收结果:阻断错误要落到具体字段上。"""
    spec = _spec()
    draft = _draft(manual={"brand": "Marine", "currency": "USD", "stock": "1"})
    mapped = generic.map_fields(draft, spec, image_map=_MAP)
    problems = generic.validate(mapped, spec)
    preview = export_writer.preview_tables(mapped, spec, violations=problems)

    statuses = {f["key"]: f["status"] for f in preview["rows"][0]["fields"]}
    assert statuses["price"] == "ERROR"
    assert statuses["sku_code"] == "OK"


def test_preview_explains_where_every_field_comes_from():
    """FE-232 的验收结果:每个字段来源清楚。"""
    spec = _spec()
    preview = export_writer.preview_tables(generic.map_fields(_draft(), spec, image_map=_MAP), spec)
    labels = {f["key"]: f["source_label"] for f in preview["header"]}
    assert labels["product_title"] == "文案"
    assert labels["garment_type"] == "已确认属性"
    assert labels["main_image"] == "已批准图片集"
    assert labels["brand"] == "运营手填"
    assert labels["category_path"] == "模板常量"


def test_csv_escapes_formula_injection():
    """导出文件会被发给外部合作方。一条 =HYPERLINK(...) 就够了。"""
    spec = _spec()
    draft = _draft(manual={
        "brand": "=cmd|'/c calc'!A1",
        "price": "1",
        "currency": "USD",
        "stock": "1",
    })
    text = export_writer.to_csv(generic.map_fields(draft, spec, image_map=_MAP), spec)
    assert "'=cmd" in text
    assert ",=cmd" not in text


def test_csv_still_writes_the_spu_row_when_there_are_no_skus():
    """只有表头的空表会让运营以为导出失败了。"""
    spec = _spec()
    product = CanonicalProduct(spu="SW-001", spu_attrs={}, skus=())
    mapped = generic.map_fields(_draft(product=product), spec, image_map=_MAP)
    text = export_writer.to_csv(mapped, spec)
    assert len(text.strip().splitlines()) == 2


def test_error_report_carries_row_numbers():
    spec = _spec()
    draft = _draft(manual={"brand": "Marine", "currency": "USD", "stock": "1"})
    mapped = generic.map_fields(draft, spec, image_map=_MAP)
    report = export_writer.errors_to_csv(generic.validate(mapped, spec))
    assert "价格" in report
    # 行号对人显示从 1 开始
    assert ",1," in report and ",2," in report


def test_xlsx_has_exactly_the_two_data_sheets():
    """校验报告不进工作簿 —— 多一个 sheet 就可能让解析器整份拒收。"""
    spec = _spec()
    mapped = generic.map_fields(_draft(), spec, image_map=_MAP)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return  # 环境没装 openpyxl 时跳过,不让纯测试套件依赖它
    import io as _io

    blob = export_writer.to_xlsx(mapped, spec, _meta())
    book = load_workbook(_io.BytesIO(blob))
    assert book.sheetnames == [export_writer.SHEET_SPU, export_writer.SHEET_SKU]
    sku_sheet = book[export_writer.SHEET_SKU]
    assert sku_sheet.max_row == 3  # 表头 + 2 行 SKU


def test_xlsx_cells_match_the_csv_cells():
    spec = _spec()
    mapped = generic.map_fields(_draft(), spec, image_map=_MAP)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    import csv as _csv
    import io as _io

    book = load_workbook(_io.BytesIO(export_writer.to_xlsx(mapped, spec, _meta())))
    csv_rows = list(_csv.reader(_io.StringIO(export_writer.to_csv(mapped, spec))))
    spu_count = len(spec.header_fields)

    sku_sheet = book[export_writer.SHEET_SKU]
    for index, row in enumerate(sku_sheet.iter_rows(min_row=2, values_only=True), start=1):
        expected = csv_rows[index][spu_count:]
        assert [c or "" for c in row] == expected


# 以下来自已删除的 test_code_review_round3.py(第三轮评审的阶段性容器文件)。
# 按被测模块归位,而不是按评审轮次分组 —— 否则同一个模块的测试会散落在多个文件里。

# ---------------------------------------------------------------- 4/5 指纹


def _fingerprint_draft(**kw):
    image_set = ImageSetSnapshot(
        image_set_id="is1", spu="SW-1", version=1,
        status=ImageSetStatus.APPROVED.value,
        items=(ImageItemSnapshot("m1", MediaRole.PRODUCT_FRONT, 0, "p", 800, 800,
                                 is_primary=True),),
    )
    product = CanonicalProduct(
        spu="SW-1",
        spu_attrs={"garment_type": AttrValue(
            field_name="garment_type", value="ONE_PIECE",
            source=AttributeSource.EXTRACTED, status=AttributeStatus.CONFIRMED,
            version_id="v1")},
    )
    base = dict(spu="SW-1", channel="SHEIN", site="BR", category_id="c1",
                product=product, image_set=image_set,
                copies={"en": CopySnapshot(copy_id="c-en", locale="en", version=1)})
    base.update(kw)
    return ListingDraftData(**base)


def test_fingerprint_covers_the_output_context():
    """评审第 4 条:改渠道、站点、类目、草稿 Schema 版本都要让草稿过期。

    这四样各自会换一套必填规则和一套列头。上一版的指纹对它们完全无感 ——
    一份在 SHEIN/BR 下校验通过的草稿,改成 GENERIC/MAIN 之后仍然
    显示「校验通过、可以导出」,而导出的是按旧规则算出来的字段。
    """
    import dataclasses

    base = _fingerprint_draft()
    stored = base.source_fingerprint()
    for field_name, value in (
        ("channel", "GENERIC"),
        ("site", "MAIN"),
        ("category_id", "c2"),
        ("schema_version", "2"),
        ("spu", "SW-2"),
    ):
        changed = dataclasses.replace(base, **{field_name: value})
        assert changed.source_fingerprint() != stored, f"改了 {field_name} 指纹却没变"


def test_rebinding_copies_to_another_locale_changes_the_fingerprint():
    """评审第 5 条:同一批文案换个语言键挂一遍,指纹必须变。

    只按文案对象值参与哈希的话,`{"en": 甲}` 和 `{"pt-BR": 甲}` 得到
    同一个指纹 —— 整批文案被绑到错的语言上,而草稿显示一切正常。
    """
    import dataclasses

    base = _fingerprint_draft()
    stored = base.source_fingerprint()
    rebound = dataclasses.replace(
        base, copies={"pt-BR": CopySnapshot(copy_id="c-en", locale="en", version=1)}
    )
    assert rebound.source_fingerprint() != stored


def test_a_mismatched_locale_key_blocks_export():
    """指纹只回答"变没变",不回答"对不对" —— 第一次就挂错的那份要另外拦。"""
    import dataclasses

    bad = dataclasses.replace(
        _fingerprint_draft(),
        copies={"pt-BR": CopySnapshot(copy_id="c-en", locale="en", version=1)},
    )
    assert copy_locale_problems(bad)
    assert not can_export(DraftStatus.VALIDATED.value, bad.source_fingerprint(), bad)
    assert copy_locale_problems(_fingerprint_draft()) == []

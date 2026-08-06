"""A41 走读修掉的几件事,各配一条能真的红的测试。

三条是普通的纯函数断言,两条是**结构性门禁**(AST)。后两条之所以只能这么写:
它们守的不变量涉及数据库会话的运行期行为,而这个套件跑不了数据库 ——
但那两个缺陷在源码结构上是可见的,所以按 `test_batch_lease.py` /
`test_workbench_query_budget.py` 的先例用 AST 钉住。
"""
from __future__ import annotations

import ast
from datetime import UTC, datetime, timedelta, timezone

from app.core.clock import iso_utc
from app.core.search import escape_like, like_pattern
from app.workbench.batch import MAX_ITEM_ATTEMPTS, ReclaimAction, reclaim_verdict
from tests.pure._helpers import BACKEND_ROOT

APP = BACKEND_ROOT / "app"


# ================================================================ LIKE 通配符


def test_underscore_is_escaped_because_real_skus_contain_it():
    # 不转义的话 `SW-001_BLK` 会命中 `SW-001XBLK` —— 而下划线在真实 SKU 里很常见
    assert like_pattern("SW-001_BLK") == "%SW-001\\_BLK%"


def test_a_bare_percent_does_not_become_match_everything():
    # 转义之前这个输入会拼成 `%%%`,也就是「返回全表」
    assert like_pattern("%") == "%\\%%"


def test_the_escape_char_itself_is_escaped_first():
    # 顺序错了的话 `\_` 会变成 `\\_`(字面反斜杠 + 仍然生效的通配符)
    assert escape_like("\\_") == "\\\\\\_"


def test_ordinary_input_is_left_alone():
    assert like_pattern("  泳装 SW-001  ") == "%泳装 SW-001%"


def test_every_ilike_call_passes_the_escape_char():
    """转义模式和 `escape=` 是一对,不许只用一半。

    只写 `like_pattern()` 而漏掉 `escape=` 的话,PostgreSQL 上仍然碰巧能工作
    (它的默认转义符正好是反斜杠),但那是实现细节 —— SQL 标准里不带 ESCAPE
    子句时转义符是未定义的。漏掉的那一处不会有任何报错。
    """
    missing: list[str] = []
    for path in APP.rglob("*.py"):
        if "__pycache__" in str(path):
            continue
        tree = ast.parse(path.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            func = node.func
            if not isinstance(func, ast.Attribute) or func.attr not in {"ilike", "like"}:
                continue
            if not any(kw.arg == "escape" for kw in node.keywords):
                missing.append(f"{path.relative_to(APP)}:{node.lineno}")
    assert not missing, f"这些 LIKE 调用没传 escape=:{missing}"


# ================================================================ 时间戳形状


def test_iso_utc_always_carries_the_offset():
    """naive 与 aware 必须序列化成**同一个**字符串。

    `expire_on_commit=False` 让「刚写完就回读」拿到 naive、「刷新页面重新查」
    拿到 aware,而裸 `.isoformat()` 会把这个差别原样透给调用方。
    """
    naive = datetime(2026, 8, 1, 10, 0, 0)
    aware = datetime(2026, 8, 1, 10, 0, 0, tzinfo=UTC)
    assert iso_utc(naive) == iso_utc(aware) == "2026-08-01T10:00:00+00:00"


def test_iso_utc_converts_other_zones_instead_of_relabelling_them():
    shanghai = datetime(2026, 8, 1, 18, 0, 0, tzinfo=timezone(timedelta(hours=8)))
    assert iso_utc(shanghai) == "2026-08-01T10:00:00+00:00"


def test_iso_utc_passes_none_through():
    assert iso_utc(None) is None


# ================================================================ 回收文案


def test_giving_up_at_the_cap_still_names_the_cap():
    verdict = reclaim_verdict(attempts=MAX_ITEM_ATTEMPTS)
    assert verdict.action is ReclaimAction.GIVE_UP
    assert f"上限 {MAX_ITEM_ATTEMPTS}" in verdict.reason


def test_after_a_human_retry_the_reason_does_not_claim_the_cap_was_blown():
    """人工重试不清 `attempts`,所以它可以大于上限。

    旧文案在那种情况下写的是「已执行 4 次仍未落地(上限 3)」—— 读起来像是
    自动重排失控了,而事实相反:那 4 次里有一次是人自己点的。
    """
    verdict = reclaim_verdict(attempts=MAX_ITEM_ATTEMPTS + 1)
    assert verdict.action is ReclaimAction.GIVE_UP
    assert "上限" not in verdict.reason
    assert str(MAX_ITEM_ATTEMPTS + 1) in verdict.reason


# ================================================================ 结构性门禁


def _function(module_path, name: str) -> ast.FunctionDef:
    tree = ast.parse(module_path.read_text(encoding="utf-8"))
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef) and node.name == name:
            return node
    raise AssertionError(f"{module_path.name} 里找不到函数 {name}")


def _assigned_attributes(node: ast.AST, target: str) -> set[str]:
    """函数体里对 `<target>.<列>` 的赋值,收成一个列名集合。"""
    out: set[str] = set()
    for sub in ast.walk(node):
        if not isinstance(sub, ast.Assign):
            continue
        for tgt in sub.targets:
            if (
                isinstance(tgt, ast.Attribute)
                and isinstance(tgt.value, ast.Name)
                and tgt.value.id == target
            ):
                out.add(tgt.attr)
    return out


def test_retry_reset_clears_the_same_columns_as_lease_reclaim():
    """「把条目打回 PENDING」只有一种正确的行结构,两处必须一致。

    `reset_items_for_retry()`(人点重试)和 `reap_expired_leases()` 的 REQUEUE
    分支(租约回收)是同一个状态转换。漏掉 `finished_at` 的后果不是脏数据:
    `_liveness_out()` 把全部行的 `max(finished_at)` 当批次心跳,于是刚点完重试
    的批次会立刻被报成 STALLED,提示「worker 中途退出了」。
    """
    path = APP / "workbench" / "batch_service.py"
    reset = _function(path, "reset_items_for_retry")
    reap = _function(path, "reap_expired_leases")

    reset_cols = _assigned_attributes(reset, "row")
    # 回收器那一头同时写 REQUEUE 与 GIVE_UP 两个分支,取 REQUEUE 那半:
    # 它是唯一一个把 status 写回 PENDING 的分支
    requeue_cols: set[str] = set()
    for branch in ast.walk(reap):
        if isinstance(branch, ast.If):
            cols = _assigned_attributes(ast.Module(body=branch.body, type_ignores=[]), "row")
            if "status" in cols and "error_code" in cols:
                requeue_cols = cols
                break

    assert requeue_cols, "没找到 reap_expired_leases 的 REQUEUE 分支"
    for column in ("status", "error_code", "error_message", "target_ref", "finished_at"):
        assert column in reset_cols, f"reset_items_for_retry 漏了 {column}"
        assert column in requeue_cols, f"reap_expired_leases 的 REQUEUE 分支漏了 {column}"


def test_read_only_endpoints_roll_back_after_the_last_orm_read():
    """只读接口的 `session.rollback()` 必须是**返回前的最后一件事**。

    `Session.rollback()` 会把会话里所有对象置为 expired,而 expired 对象的
    下一次属性访问会隐式发一条 SELECT。放在出参组装之前的话,后面每一处
    `ctx.product.<列>` 都是一次往返 —— 最贵的是 `sort_in_memory` 的 tiebreak,
    它跑在全量结果上而不是当前页。

    `test_workbench_query_budget.py` 结构上看不见这一类:它数的是源码里的
    `session.*` 调用点,而这里一个都没有。所以在这里按位置钉。
    """
    path = APP / "api" / "workbench.py"
    for name in ("list_products", "product_flow"):
        fn = _function(path, name)
        found = [
            i
            for i, stmt in enumerate(fn.body)
            for sub in ast.walk(stmt)
            if isinstance(sub, ast.Call)
            and isinstance(sub.func, ast.Attribute)
            and sub.func.attr == "rollback"
        ]
        # 恰好一次:补一次"早一点的保险"就等于把 expire 提前,后面所有 ORM
        # 读又变回逐行往返 —— 而末尾那一次仍然在,位置断言看不出来
        assert len(found) == 1, f"{name} 应当**只有一次**只读兜底回滚,实际 {len(found)} 次"
        tail = fn.body[found[0] + 1 :]
        assert len(tail) == 1 and isinstance(tail[0], ast.Return), (
            f"{name}:rollback 之后只允许剩一条 return,"
            f"实际还有 {len(tail)} 条语句 —— 它们会各自触发一次隐式 SELECT"
        )
        assert isinstance(tail[0].value, ast.Name), (
            f"{name}:返回值必须是回滚前就组装好的变量,不能在 return 里现读 ORM 对象"
        )


# ================================================================ 批量导出的归属列


def _batch_book():
    """两个 SPU 的批量工作簿。openpyxl 缺席时返回 None,由调用方跳过。"""
    from app.channels import generic
    from app.listings import export_writer
    from tests.pure.test_listing_mapping_export import _draft, _meta, _spec

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, None
    import io

    spec = _spec()
    first = generic.map_fields(_draft(), spec)
    second = generic.map_fields(_draft(spu="SW-002"), spec)
    blob = export_writer.to_batch_xlsx(
        [
            export_writer.BatchEntry(spu="SW-001", mapped=first, sku_count=len(first.rows)),
            export_writer.BatchEntry(spu="SW-002", mapped=second, sku_count=len(second.rows)),
        ],
        spec,
        _meta(),
    )
    return load_workbook(io.BytesIO(blob)), spec


def test_batch_sku_rows_can_name_their_parent_spu():
    """批量 SKU 表必须能接回 SPU 表。

    单件导出里这个关系是隐式的(SPU 表就一行)。批量把 N 个 SPU 的 SKU 行拼进
    同一张表,而 `row_fields` 里没有任何字段指向父 SPU —— 少了归属列,400 行
    SKU 就没有一列能说明某一行属于谁,`manifest.csv` 也补不上(它的粒度在 SPU,
    而且平台拿不到它)。
    """
    from app.listings import export_writer

    book, spec = _batch_book()
    if book is None:
        return

    sheet = book[export_writer.SHEET_SKU]
    headers = [c.value for c in sheet[1]]
    assert headers[0] == export_writer.spu_link_header(spec), (
        f"批量 SKU 表第一列应当是归属列,实际是 {headers[0]!r}"
    )

    owners = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert owners == {"SW-001", "SW-002"}, f"归属列没有覆盖全部 SPU:{owners}"


def test_batch_link_column_reuses_the_spu_sheet_header():
    """归属列和 SPU 表的 SPU 列是同一个词,不另起名字。

    平台的导入器按列头定位(见 `_headers`)。用同名列它至少有机会认出来,
    另造一个「所属SPU」它一定认不出。
    """
    from app.listings import export_writer

    book, spec = _batch_book()
    if book is None:
        return

    spu_headers = [c.value for c in book[export_writer.SHEET_SPU][1]]
    link = export_writer.spu_link_header(spec)
    assert link in spu_headers, f"归属列 {link!r} 在 SPU 表里找不到同名列"


def test_single_export_shape_is_untouched():
    """单件导出不许因为批量的修法而多出一列 —— 它没有这个问题。"""
    from app.channels import generic
    from app.listings import export_writer
    from tests.pure.test_listing_mapping_export import _draft, _meta, _spec

    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    import io

    spec = _spec()
    mapped = generic.map_fields(_draft(), spec)
    book = load_workbook(io.BytesIO(export_writer.to_xlsx(mapped, spec, _meta())))
    headers = [c.value for c in book[export_writer.SHEET_SKU][1]]
    assert headers == [f.column_header or f.key for f in spec.row_fields]


# ================================================================ 清理计划的数量口径


def test_plan_counts_match_what_delist_will_actually_queue():
    """给人看的数字必须是实际会执行的数量。

    `occupying` / `not_delistable` 只数真实商品(`for r in real`),而
    `run_delist` 遍历的是 `plan.rows` —— **全部行,含 SIM-**。人工测试批次里
    几乎全是模拟行,两个口径能差一个数量级,而这个模块的整套安全设计就是
    「强制限定作用域 + 动手前把数量给人看」。

    `cleanup_service` 要 import sqlalchemy,纯测试套件里跑不起来(同
    `test_delist_and_cleanup.py` 开头的说明),所以这里用 AST 钉结构:
    `to_queue` / `to_skip` 的推导式必须遍历 `rows` 而不是 `real`。
    """
    src = (APP / "services" / "cleanup_service.py").read_text(encoding="utf-8")
    report_fn = _function_in(src, "_report")

    iterated: dict[str, str] = {}
    for kw in next(
        node for node in ast.walk(report_fn)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Name)
        and node.func.id == "CleanupReport"
    ).keywords:
        if kw.arg not in {"occupying", "not_delistable", "to_queue", "to_skip"}:
            continue
        comps = [n for n in ast.walk(kw.value) if isinstance(n, ast.comprehension)]
        assert comps, f"{kw.arg} 不再是推导式,这条断言要跟着改"
        iterated[kw.arg] = ast.unparse(comps[0].iter)

    assert iterated["occupying"] == "real", "占位数应当只算真实商品"
    assert iterated["not_delistable"] == "real", "不可下架数应当只算真实商品"
    # **执行口径必须遍历全部行** —— run_delist 就是这么遍历的
    assert iterated["to_queue"] == "rows", (
        f"to_queue 遍历的是 {iterated['to_queue']},而 run_delist 遍历 plan.rows"
    )
    assert iterated["to_skip"] == "rows", (
        f"to_skip 遍历的是 {iterated['to_skip']},而 run_delist 遍历 plan.rows"
    )


def test_the_cleanup_script_prints_the_execution_count():
    """脚本必须把执行口径打出来。

    报表里有这个数、而 `--apply` 之前打给人看的那几行没有,等于没修:
    运营看的是 stderr 那几行,不是 JSON。
    """
    src = (APP / "scripts" / "cleanup_test_listings.py").read_text(encoding="utf-8")
    summary = _function_in(src, "_summary")
    printed = ast.unparse(summary)
    assert "to_queue" in printed and "to_skip" in printed, (
        "_summary 没有打印执行口径的数量"
    )
    _assert_reachable(summary, where="_summary 的执行口径那几行")


def _assert_reachable(node: ast.AST, *, where: str) -> None:
    """这段代码不能被一个恒假条件绕过去。

    存在的理由很具体:上一轮给这几条测试做变异验证时,把守卫改成
    `if False and _delivery_pending(...)`、把打印改成 `if False:`,
    两条测试**都没红** —— 它们只查了文本里有没有那个名字,而名字还在。

    只查文本的结构测试就是这样失效的:它守的是"有没有写",不是"生不生效"。
    """
    for sub in ast.walk(node):
        if isinstance(sub, (ast.If, ast.IfExp)):
            for part in ast.walk(sub.test):
                if isinstance(part, ast.Constant) and part.value is False:
                    raise AssertionError(f"{where}被一个恒假条件绕过去了")


def _function_in(src: str, name: str) -> ast.FunctionDef:
    for node in ast.walk(ast.parse(src)):
        if isinstance(node, ast.FunctionDef) and node.name == name:
            return node
    raise AssertionError(f"找不到函数 {name}")


# ================================================================ 轮询不越权


def test_polling_defers_to_delivery_while_an_outbox_row_is_open():
    """投递没落地时,轮询不许改本地状态。

    `publish_policy._status_for` 早就写下了这件事的一半 —— DELIST 成功时平台
    通常仍然返回那条商品记录、带着 `LISTED` 字样,照着翻译会把一次成功的下架
    记成「在架」。投递路径能防住是因为它知道自己刚发的是什么操作;轮询路径
    不知道,`classify_poll` 只拿到状态码和 body。

    而 `DELISTING` / `UPDATING` / `REPUBLISHING` 全在 `POLLABLE_STATUSES` 里,
    所以「运营点下架 -> 轮询节拍先到 -> 平台说 LISTED -> 状态翻回在架 ->
    LISTED 不可轮询,这一行掉出轮询集合」是常态而不是边角。

    `poll_service` 要 import sqlalchemy,纯套件里跑不起来,所以用 AST 钉:
    赋值 `listing.status` 之前必须先问过 `_delivery_pending`。
    """
    src = (APP / "services" / "poll_service.py").read_text(encoding="utf-8")
    fn = _function_in(src, "apply_poll_outcome")

    body = fn.body
    guard_at = None
    assign_at = None
    for index, stmt in enumerate(body):
        dumped = ast.unparse(stmt)
        if guard_at is None and "_delivery_pending" in dumped:
            guard_at = index
        if assign_at is None and "listing.status = outcome.listing_status" in dumped:
            assign_at = index

    assert guard_at is not None, "apply_poll_outcome 里没有 _delivery_pending 守卫"
    assert assign_at is not None, "找不到写 listing.status 的语句,这条断言要跟着改"
    assert guard_at < assign_at, (
        "守卫排在写状态之后 —— 那等于没有守卫"
    )
    _assert_reachable(body[guard_at], where="apply_poll_outcome 的投递守卫")


def test_the_guard_still_lets_a_platform_rejection_through():
    """平台明确拒了不算「你的请求还没处理完」,必须放行。

    挡住它的话 `_reflow_rejection` 不会跑,驳回记不进台账,商品会顶着
    「下架中 / 更新中」沉底,而它其实已经被平台拒了。
    """
    src = (APP / "services" / "poll_service.py").read_text(encoding="utf-8")
    fn = _function_in(src, "apply_poll_outcome")
    guard = next(
        stmt for stmt in fn.body
        if "_delivery_pending" in ast.unparse(stmt)
    )
    assert "PLATFORM_REJECTED" in ast.unparse(guard), (
        "守卫没有给 PLATFORM_REJECTED 留出口"
    )


def test_the_guard_is_self_limiting_not_a_status_whitelist():
    """判据必须是「有没有未结算的 outbox 行」,不是「状态在不在某个集合里」。

    写成状态集合的话,投递以 DEAD 收场且没写回 listing 状态时,这一行会被
    永久钉在 DELISTING —— 轮询再也改不动它,而 outbox 里已经没有人会来救。
    查 outbox 是自限的:一结算,轮询立刻恢复全部权威。
    """
    src = (APP / "services" / "poll_service.py").read_text(encoding="utf-8")
    fn = _function_in(src, "_delivery_pending")
    # 只看代码,不看文档字符串 —— 那里正解释着 DONE / DEAD 为什么**不**在里面
    code = [
        s
        for s in fn.body
        if not (isinstance(s, ast.Expr) and isinstance(s.value, ast.Constant))
    ]
    dumped = "\n".join(ast.unparse(s) for s in code)
    assert "PublishOutbox" in dumped and "PublishAttempt" in dumped, (
        "_delivery_pending 没有去查 outbox"
    )
    for settled in ("PENDING", "LEASED"):
        assert settled in dumped, f"未结算状态漏了 {settled}"
    assert "DONE" not in dumped and "DEAD" not in dumped, (
        "DONE / DEAD 是结算过的,不该出现在未结算判据里"
    )

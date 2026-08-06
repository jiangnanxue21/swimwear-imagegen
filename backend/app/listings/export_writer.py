"""上架文件导出(任务书 FE-235 / BE-204)。**零数据库依赖。**

## 一个不变量

阶段 2 的退出指标里有一条「导出一致率 100%:Excel/CSV 字段与页面预览一致」。
保证它只有一个办法:**页面预览和导出文件读的是同一份 `mapped_payload`**。

于是这个模块的入参是一份已经映射好的 `MappedListing`,而不是商品 id ——
它没有能力去查一次库、拿到一份和预览不一样的数据。想让两者不一致,
得先把这个函数签名改掉,而那是一次会出现在 diff 里的动作。

## 三种格式的分工

    xlsx   给运营。两个 sheet,和页面上两张预览表逐格对应
    csv    给表格工具与外部合作方。扁平化:SPU 列在每个 SKU 行上重复
    json   给程序。原样输出 header/rows,不做任何形状变换

CSV 的扁平化是刻意的:一个文件只能有一张表,而运营拿到分成两段的 CSV
第一件事一定是手工把 SPU 那几列贴到每一行上。这里直接给出他要用的形状,
**值本身与预览逐格相同**。
"""
from __future__ import annotations

import csv
import io
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any

from app.channels.base import ChannelFieldSpec
from app.listings.contracts import FieldViolation, MappedListing
from app.services.export_format import FORMULA_PREFIXES, escape_formula

#: 工作簿里的 sheet 名。和页面上两张预览表的标题保持一致 ——
#: 运营对照检查时不需要做一次心算映射。
SHEET_SPU = "商品信息"
SHEET_SKU = "SKU明细"

#: 校验报告是**独立文件**,不往工作簿里加 sheet(§9.2 意见 #13)。
#: 平台的解析器可能按 sheet 序号读,多一个 sheet 就可能整份被拒。
#: 本模板是我们自己的,但这个习惯值得保持:导出的表就该只有数据。
VALIDATION_SHEET_FORBIDDEN = True


@dataclass(frozen=True)
class ExportMeta:
    """导出时要带上的身份信息。进文件名,也进 JSON。"""

    spu: str
    channel: str
    site: str
    category_id: str
    spec_version: str
    draft_id: str | None = None
    fingerprint: str | None = None
    exported_at: datetime | None = None
    exported_by: str | None = None


def _headers(spec: ChannelFieldSpec) -> tuple[list[str], list[str], list[str], list[str]]:
    """(SPU 列头, SPU 字段键, SKU 列头, SKU 字段键)。

    列头与键分开返回,因为文件里写的是列头,而定位单元格用的是键。
    Excel 按列头定位、不按列序号 —— 平台加一列就会让所有按序号写的
    代码错位,而且错得很安静。
    """
    return (
        [f.column_header or f.key for f in spec.header_fields],
        [f.key for f in spec.header_fields],
        [f.column_header or f.key for f in spec.row_fields],
        [f.key for f in spec.row_fields],
    )


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    return str(value)


def _force_text_cells(sheet, first_row: int = 1) -> None:
    """把整张表的单元格钉成文本类型。

    ## 这里修的是一个真实缺口

    `to_csv` 每一格都过 `escape_formula`,而 xlsx 这条路径上没有对应的处理。
    openpyxl 会**按内容推断类型**:赋一个以 `=` 开头的字符串,它写出来的是
    公式(`data_type='f'`),不是文本。于是一个叫 `=UPF50+ 防晒款` 的标题
    在 Excel 里变成一条公式,显示 `#NAME?`,而页面预览显示的是原文 ——
    §5.3「导出一致率 100%」在这一格上不成立,平台拿到的也是错的值。

    (`+`、`-`、`@` 开头的字符串 openpyxl 仍按文本处理,只有 `=` 会被转成
    公式;但这里不区分,统一钉成文本 —— 少一条要记住的例外。)

    ## 为什么不复用 CSV 那套加单引号的做法

    CSV 是纯文本,前导单引号是**给 Excel 解析器看的指令**,不进单元格内容。
    xlsx 里没有这层解析:写进去的 `'=x` 就是三个字符,Excel 会原样显示那个
    引号 —— 值就真的变了,反而把一致性从"可能不一致"变成"一定不一致"。

    xlsx 的正确做法是设 `data_type='s'`(明确声明这格是字符串)加上
    `quotePrefix`(Excel 用来标记"这是文本"的样式位)。值本身一个字符不动。
    """
    for row in sheet.iter_rows(min_row=first_row):
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"
                if cell.value.startswith(FORMULA_PREFIXES):
                    cell.quotePrefix = True


# ---------------------------------------------------------------- JSON


def to_json(
    mapped: MappedListing, spec: ChannelFieldSpec, meta: ExportMeta
) -> dict[str, Any]:
    """给程序用。原样输出,不做形状变换。"""
    spu_labels, spu_keys, sku_labels, sku_keys = _headers(spec)
    return {
        "meta": {
            "spu": meta.spu,
            "channel": meta.channel,
            "site": meta.site,
            "category_id": meta.category_id,
            "spec_version": meta.spec_version,
            "draft_id": meta.draft_id,
            "source_fingerprint": meta.fingerprint,
            "exported_at": (meta.exported_at or datetime.now(UTC)).isoformat(),
            "exported_by": meta.exported_by,
        },
        "columns": {
            "header": [
                {"key": k, "label": lbl} for k, lbl in zip(spu_keys, spu_labels, strict=True)
            ],
            "rows": [{"key": k, "label": lbl} for k, lbl in zip(sku_keys, sku_labels, strict=True)],
        },
        "header": {k: mapped.header.get(k) for k in spu_keys},
        "rows": [{k: row.get(k) for k in sku_keys} for row in mapped.rows],
    }


# ---------------------------------------------------------------- CSV


def to_csv(mapped: MappedListing, spec: ChannelFieldSpec) -> str:
    """扁平化 CSV:SPU 列在每个 SKU 行上重复。

    CRLF + 调用方的 utf-8-sig,让 Excel 双击打开不乱码、不错行。
    每个单元格都过一次公式转义 —— 这个文件会被发给外部合作方,
    一条 `=HYPERLINK(...)` 就够了。
    """
    spu_labels, spu_keys, sku_labels, sku_keys = _headers(spec)

    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow([*spu_labels, *sku_labels])

    spu_cells = [escape_formula(_cell(mapped.header.get(k))) for k in spu_keys]
    if not mapped.rows:
        # 一行 SKU 都没有时仍然写出 SPU 那一段。
        # 写一张只有表头的空表,运营会以为导出失败了
        writer.writerow([*spu_cells, *["" for _ in sku_keys]])
    for row in mapped.rows:
        writer.writerow(
            [*spu_cells, *[escape_formula(_cell(row.get(k))) for k in sku_keys]]
        )
    return buffer.getvalue()


def errors_to_csv(violations: Sequence[FieldViolation]) -> str:
    """校验报告(独立文件)。

    有它才谈得上「每个失败商品均可跳转到具体问题位置」—— 报告里
    带行号和字段键,而不是一句「校验未通过」。
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(["级别", "字段", "行号", "问题码", "说明"])
    for v in violations:
        writer.writerow(
            [
                "阻断" if v.is_blocking else "提醒",
                escape_formula(v.field_key),
                "" if v.row_index is None else str(v.row_index + 1),
                escape_formula(v.code),
                escape_formula(v.message),
            ]
        )
    return buffer.getvalue()


# ---------------------------------------------------------------- Excel


def to_xlsx(
    mapped: MappedListing, spec: ChannelFieldSpec, meta: ExportMeta
) -> bytes:
    """生成 Excel 工作簿。两个 sheet,和页面上两张预览表逐格对应。

    openpyxl 是延迟导入的:导出是一条不常走的路径,而 `app.listings`
    会被大量纯函数测试导入 —— 让它们为了一个不用的依赖装一个包不合理。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - 环境问题,不是逻辑问题
        from app.core.errors import ConfigError

        raise ConfigError(
            "导出 Excel 需要 openpyxl,当前环境没有安装。"
            "先 pip install openpyxl,或改用 CSV 导出"
        ) from exc

    spu_labels, spu_keys, sku_labels, sku_keys = _headers(spec)

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B4B5A")  # 与前端主色同一个值
    wrap = Alignment(vertical="top", wrap_text=True)

    def write_sheet(sheet, labels: list[str], rows: list[list[Any]]) -> None:
        sheet.append(labels)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for index, label in enumerate(labels, start=1):
            width = max(12, min(48, len(str(label)) * 2 + 6))
            sheet.column_dimensions[get_column_letter(index)].width = width
        # 必须在 append 之后:openpyxl 是在赋值那一刻推断类型的
        _force_text_cells(sheet)

    spu_sheet = workbook.active
    spu_sheet.title = SHEET_SPU
    write_sheet(
        spu_sheet, spu_labels, [[_cell(mapped.header.get(k)) for k in spu_keys]]
    )

    sku_sheet = workbook.create_sheet(SHEET_SKU)
    write_sheet(
        sku_sheet,
        sku_labels,
        [[_cell(row.get(k)) for k in sku_keys] for row in mapped.rows],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------- 预览


def preview_tables(
    mapped: MappedListing,
    spec: ChannelFieldSpec,
    *,
    violations: Sequence[FieldViolation] = (),
) -> dict[str, Any]:
    """页面预览用的结构(FE-232)。

    **和导出读同一份 `mapped`。** 每个字段带上来源、状态和问题,
    于是运营在页面上看到的每一格,都能回答「这个值是哪来的」和
    「它为什么是红的」。
    """
    from app.channels.generic import source_label

    by_field: dict[tuple[str, int | None], list[FieldViolation]] = {}
    for v in violations:
        by_field.setdefault((v.field_key, v.row_index), []).append(v)

    def describe(field, value: Any, row_index: int | None) -> dict[str, Any]:
        problems = by_field.get((field.key, row_index), [])
        blocking = [p for p in problems if p.is_blocking]
        return {
            "key": field.key,
            "label": field.column_header or field.key,
            "value": None if value is None else _cell(value),
            "required": field.required,
            "source": field.source,
            "source_label": source_label(field),
            "max_length": field.max_length,
            "status": "ERROR" if blocking else ("WARNING" if problems else "OK"),
            "problems": [
                {"code": p.code, "message": p.message, "level": p.level}
                for p in problems
            ],
        }

    return {
        "header": [
            describe(f, mapped.header.get(f.key), None) for f in spec.header_fields
        ],
        "rows": [
            {
                "row_index": index,
                "sku": row.get("sku_code"),
                "fields": [describe(f, row.get(f.key), index) for f in spec.row_fields],
            }
            for index, row in enumerate(mapped.rows)
        ],
    }


def field_source_map(spec: ChannelFieldSpec) -> Mapping[str, str]:
    """字段键 -> 来源表达式。审计与排错时用。"""
    return {
        f.key: (f.source or "")
        for f in (*spec.header_fields, *spec.row_fields)
    }


# ---------------------------------------------------------------- 批量导出(阶段 3)


@dataclass(frozen=True)
class BatchEntry:
    """批量导出里的一件商品(一个 SPU)。"""

    spu: str
    mapped: MappedListing
    draft_id: str | None = None
    fingerprint: str | None = None
    #: 这个 SPU 下参与本次导出的 SKU 数,写进清单便于核对行数
    sku_count: int = 0


def spu_link_header(spec: ChannelFieldSpec) -> str:
    """批量 SKU 表那一列的列头。

    取 SPU 表里 `source` 是 `product.spu` 的那个字段的列头(泳装模板里是
    「SPU编码」),而不是另起一个名字:两张表指的是同一件东西,列头也该是
    同一个词。平台的导入器按列头定位(见 `_headers`),用同名列它至少有
    机会认出来;另起一个「所属SPU」它一定认不出。

    找不到就退回字面量。这一列的价值是「能把两张表接起来」,退化的列头
    仍然做得到 —— 比因为模板换了字段名就整列消失好。
    """
    for field in spec.header_fields:
        if (field.source or "").strip() == "product.spu":
            return field.column_header or field.key
    return "SPU编码"


def to_batch_xlsx(
    entries: Sequence[BatchEntry], spec: ChannelFieldSpec, meta: ExportMeta
) -> bytes:
    """把多个 SPU 写进一个工作簿(阶段 3 FE-301 批量导出)。

    ## 和单件导出的唯一形状差异:SKU 表多一列「所属 SPU」

    其余完全一致(两个 sheet、同样的表头、同样的取值来源),理由是
    「导出一致率 100%:Excel 字段与页面预览一致」是阶段 2 的退出指标,
    批量路径换一种形状那条指标就在运营真正用的路径上失效了。

    但那一列必须加。单件导出里两张表的对应关系是**隐式**的 —— SPU 表就
    一行,SKU 表全是它的,不需要任何列来说明归属。批量导出把 N 个 SPU 的
    SKU 行拼进同一张表,这个前提当场失效:`swimwear.yaml` 的 `row_fields`
    是 `sku_code / color / size / price / currency / stock / lead_time_days /
    variant_image`,**没有一个字段指向父 SPU**。50 个 SPU、400 行 SKU 拼在
    一起,没有任何一列能说明某一行属于谁。

    `manifest.csv` 补不上这个:它每个 entry 一行(批次 -> SPU -> 草稿指纹),
    粒度在 SPU 不在 SKU 行,而且平台拿不到它。所以「批次与商品的对应关系
    进独立清单文件」这句话只对**批次层**成立;**行层**的归属必须在工作簿里。

    A41 之前这一列不存在,导出的批量文件是接不起来的。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - 环境问题,不是逻辑问题
        from app.core.errors import ConfigError

        raise ConfigError(
            "导出 Excel 需要 openpyxl,当前环境没有安装。"
            "先 pip install openpyxl,或改用 CSV 导出"
        ) from exc

    spu_labels, spu_keys, sku_labels, sku_keys = _headers(spec)
    link_header = spu_link_header(spec)

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B4B5A")
    wrap = Alignment(vertical="top", wrap_text=True)

    def write_sheet(sheet, labels: list[str], rows: list[list[Any]]) -> None:
        sheet.append(labels)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for index, label in enumerate(labels, start=1):
            width = max(12, min(48, len(str(label)) * 2 + 6))
            sheet.column_dimensions[get_column_letter(index)].width = width
        # 必须在 append 之后:openpyxl 是在赋值那一刻推断类型的
        _force_text_cells(sheet)

    spu_sheet = workbook.active
    spu_sheet.title = SHEET_SPU
    write_sheet(
        spu_sheet,
        spu_labels,
        [[_cell(e.mapped.header.get(k)) for k in spu_keys] for e in entries],
    )

    sku_sheet = workbook.create_sheet(SHEET_SKU)
    write_sheet(
        sku_sheet,
        # 归属列放**第一列**:人打开文件先看到的就是「这行属于谁」,
        # 而不是滚到最右边去找
        [link_header, *sku_labels],
        [
            [_cell(entry.spu), *[_cell(row.get(k)) for k in sku_keys]]
            for entry in entries
            for row in entry.mapped.rows
        ],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def manifest_csv(
    entries: Sequence[BatchEntry], meta: ExportMeta, *, batch_id: str
) -> str:
    """导出清单。**批次 -> 商品 -> 草稿指纹**的对照表。

    这份文件回答的是出事之后的那个问题:"上周三导的那批里,这个 SPU 用的是
    哪一版草稿?"。指纹在这里是全长的(不像审计 payload 里截了 12 位),
    因为它是唯一能把文件里的一行和库里的一版草稿对上的东西。
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(
        [
            "批次号",
            "SPU",
            "SKU数",
            "草稿ID",
            "草稿指纹",
            "模板版本",
            "渠道",
            "站点",
            "导出时间",
            "导出人",
        ]
    )
    exported_at = (meta.exported_at or datetime.now(UTC)).isoformat()
    for entry in entries:
        writer.writerow(
            [
                escape_formula(batch_id),
                escape_formula(entry.spu),
                entry.sku_count or len(entry.mapped.rows),
                escape_formula(entry.draft_id or ""),
                escape_formula(entry.fingerprint or ""),
                escape_formula(meta.spec_version),
                escape_formula(meta.channel),
                escape_formula(meta.site),
                exported_at,
                escape_formula(meta.exported_by or ""),
            ]
        )
    return buffer.getvalue()

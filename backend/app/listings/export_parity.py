"""导出一致性比对(任务书 §5.3「导出一致率 100%」)。**判定部分零依赖纯函数。**

## 为什么需要一个脚本,而不是让人核对

§5.3 的判定方法原文写着"逐字段比对全部必填字段 + 随机 5 个非必填字段;
**建议写一次性比对脚本,人工逐字段核对 10 件商品不现实**"。而 §9.2 的
签字第 5 条要求"导出内容与页面预览一致(**比对记录留档**)"。

两句话合起来是一个明确的交付物:一个能跑出可归档结论的比对器。
在它存在之前,那条指标只能靠"看了一下,好像一样"通过 —— 而这正是
V1.1 通篇在防的那种验收。

## 这个比对器能证明什么、不能证明什么

**能证明的**,是"页面上那一格"和"文件里那一格"是同一个字符串。
这条链路上真实存在几个会让两者分叉的地方:

    openpyxl 的类型推断    以 `=` 开头的字符串会被写成**公式**
    None 与空串            预览给 `null`,文件里是空单元格
    列头与列序             预览按 spec 顺序渲染,文件按 append 顺序写
    批量导出               `to_batch_xlsx` 是另一个函数,不是同一段代码
    行数                   一个 SKU 丢了,总数对不上但每一格都对

**不能证明的**,是映射本身对不对。两边都读同一份 `mapped_payload`,
所以映射错了会一致地错 —— 那是 §5.2 第 11 步人工核对与预览里的
`source` 列负责的事。把这件事写清楚,是为了避免有人拿一份全绿的
比对报告去回答"字段映射对不对"这个问题。

## 为什么从预览出发,而不是从 `mapped` 出发

拿 `to_xlsx(mapped)` 和 `mapped` 自己比是循环论证:两边同源,永远相等,
测试永远绿,而绿得没有信息。所以期望值取自 `preview_tables()` 的出参 ——
**运营眼睛看到的那个结构**,包含列头文案、字段顺序和渲染后的值。
"""
from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Any


#: 单元格值的规范化规则。**只有这一条**,而且两边都过它。
#:
#: 预览里 `value` 为 `None` 表示"这个字段没值";Excel 里同一件事是空单元格,
#: openpyxl 读回来是 `None`。它们语义相同、表示不同,不归一化会产出一堆
#: 假不一致,而假不一致的代价是这个脚本第一次跑就被当成噪音关掉。
#:
#: 除此之外不做任何清洗 —— 不 strip 空格、不统一大小写、不折叠换行。
#: 那些差异是真差异:平台按字符数截断,一个尾随空格就能让标题超长。
def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


@dataclass(frozen=True)
class SheetTable:
    """一张表。可以来自真实文件,也可以来自预览出参。"""

    name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]

    @property
    def cell_count(self) -> int:
        return sum(len(r) for r in self.rows)


#: 不一致的种类。分类不是装饰:HEADER_ORDER 和 CELL 的修法完全不同,
#: 混成一句"不一致"会让人从第一格开始逐格看。
KIND_LABELS: Mapping[str, str] = {
    "MISSING_SHEET": "文件里缺这张表",
    "EXTRA_SHEET": "文件里多了一张表",
    "HEADER_SET": "列集合不一致",
    "HEADER_ORDER": "列顺序不一致",
    "ROW_COUNT": "行数不一致",
    "CELL": "单元格值不一致",
}


@dataclass(frozen=True)
class Mismatch:
    kind: str
    sheet: str
    message: str
    #: 1 起的数据行号(不含表头那一行),表级问题为 None
    row: int | None = None
    column: str | None = None
    expected: str | None = None
    actual: str | None = None

    @property
    def kind_label(self) -> str:
        return KIND_LABELS.get(self.kind, self.kind)

    def as_row(self) -> list[str]:
        """留档用的一行。列顺序与 `REPORT_HEADER` 对应。"""
        return [
            self.sheet,
            self.kind_label,
            "" if self.row is None else str(self.row),
            self.column or "",
            "" if self.expected is None else self.expected,
            "" if self.actual is None else self.actual,
            self.message,
        ]


#: 比对报告的列头(§9.2「比对记录留档」)。
REPORT_HEADER: tuple[str, ...] = (
    "表",
    "问题",
    "行",
    "列",
    "页面预览值",
    "文件实际值",
    "说明",
)


@dataclass(frozen=True)
class ParityReport:
    mismatches: tuple[Mismatch, ...] = ()
    checked_cells: int = 0
    #: 比对了哪些表,给报告用
    sheets: tuple[str, ...] = ()

    @property
    def ok(self) -> bool:
        return not self.mismatches

    @property
    def rate(self) -> float:
        """一致率。分母是**比对过的单元格数**,不是表格总面积。

        §5.3 要的是 100%,所以这个数字的用途不是"达到 95% 就行",
        而是让不一致的规模一眼可见:1 格错和 300 格错是两种不同的故障。
        """
        cell_problems = sum(1 for m in self.mismatches if m.kind == "CELL")
        if self.checked_cells <= 0:
            return 1.0 if self.ok else 0.0
        return max(0.0, 1.0 - cell_problems / self.checked_cells)

    def summary(self) -> str:
        if self.ok:
            return f"一致:比对 {self.checked_cells} 格,全部相同"
        by_kind: dict[str, int] = {}
        for m in self.mismatches:
            by_kind[m.kind] = by_kind.get(m.kind, 0) + 1
        parts = ", ".join(
            f"{KIND_LABELS.get(k, k)} {n} 处" for k, n in sorted(by_kind.items())
        )
        return f"不一致:比对 {self.checked_cells} 格,{parts}(一致率 {self.rate:.2%})"


# ==========================================================
# 期望值:从页面预览出参还原出「文件应该长什么样」
# ==========================================================


def expected_tables(
    preview: Mapping[str, Any], *, spu_sheet: str, sku_sheet: str
) -> dict[str, SheetTable]:
    """把 `export_writer.preview_tables()` 的出参翻成两张期望表。

    预览的结构是 `{"header": [field...], "rows": [{"fields": [field...]}]}`,
    每个 field 带 `label` 与 `value`。文件里写的是 label 那一行加上 value 那些行,
    所以这个函数就是把预览"压平"成表格形状 —— **不碰任何值**。
    """
    header_fields = list(preview.get("header") or ())
    spu = SheetTable(
        name=spu_sheet,
        headers=tuple(str(f.get("label") or f.get("key") or "") for f in header_fields),
        rows=(tuple(normalize(f.get("value")) for f in header_fields),),
    )

    row_entries = list(preview.get("rows") or ())
    sku_headers: tuple[str, ...] = ()
    if row_entries:
        first = list(row_entries[0].get("fields") or ())
        sku_headers = tuple(
            str(f.get("label") or f.get("key") or "") for f in first
        )
    sku = SheetTable(
        name=sku_sheet,
        headers=sku_headers,
        rows=tuple(
            tuple(normalize(f.get("value")) for f in (entry.get("fields") or ()))
            for entry in row_entries
        ),
    )
    return {spu_sheet: spu, sku_sheet: sku}


def merge_expected(
    tables: Sequence[Mapping[str, SheetTable]],
    *,
    spu_sheet: str,
    sku_sheet: str,
    link_header: str | None = None,
    spus: Sequence[str] = (),
) -> dict[str, SheetTable]:
    """多个商品的期望表拼成一份(批量导出用)。

    批量导出的形状是"同样两个 sheet,只是行数变多"—— **除了 SKU 表多一列
    「所属 SPU」**(见 `export_writer.to_batch_xlsx`)。所以拼法是按顺序接行,
    列头取第一个非空的那一份;如果某个商品的列头与别人不同,那本身就是一处
    不一致,交给 `compare` 去报,这里不消化掉它。

    `link_header` + `spus` 一起传时,给 SKU 表补上那一列 —— 页面预览是单件的,
    它当然没有这一列,而文件里有。不补的话这个比对会把**批量导出唯一正确的
    那处形状差异**报成列头错位,于是每次跑都红,然后没人再跑它。

    `spus` 与 `tables` 按下标对齐(调用方在同一个循环里 append 两者)。
    对不齐是编码错误,不是数据问题,所以直接断言。
    """
    if link_header is not None and len(spus) != len(tables):
        raise ValueError(
            f"spus({len(spus)}) 与 tables({len(tables)}) 数量对不上,"
            "批量归属列会错位"
        )
    merged: dict[str, SheetTable] = {}
    for sheet_name in (spu_sheet, sku_sheet):
        headers: tuple[str, ...] = ()
        rows: list[tuple[str, ...]] = []
        add_link = link_header is not None and sheet_name == sku_sheet
        for index, one in enumerate(tables):
            table = one.get(sheet_name)
            if table is None:
                continue
            if not headers and table.headers:
                headers = (
                    (str(link_header), *table.headers) if add_link else table.headers
                )
            if add_link:
                spu = str(spus[index])
                rows.extend((spu, *row) for row in table.rows)
            else:
                rows.extend(table.rows)
        merged[sheet_name] = SheetTable(
            name=sheet_name, headers=headers, rows=tuple(rows)
        )
    return merged


# ==========================================================
# 比对
# ==========================================================


def compare(
    expected: Mapping[str, SheetTable],
    actual: Mapping[str, SheetTable],
    *,
    max_cell_mismatches: int = 200,
) -> ParityReport:
    """逐格比对。**不提前 return** —— 一份报告要能一次说完所有问题。

    `max_cell_mismatches` 只截断单元格级别的条目:列头错位会让每一格都不同,
    截断之后仍然会有一条 HEADER_ORDER 说明根因。不截断的话,一次列错位
    能刷出上万行报告,而运营需要的只是"列错位了"这一句。
    """
    mismatches: list[Mismatch] = []
    checked = 0
    cell_problems = 0

    for name in sorted(set(actual) - set(expected)):
        mismatches.append(
            Mismatch(
                kind="EXTRA_SHEET",
                sheet=name,
                message=(
                    "文件里有页面预览没有的表。平台解析器可能按 sheet 序号读,"
                    "多一张表就可能整份被拒"
                ),
            )
        )

    for name, want in expected.items():
        got = actual.get(name)
        if got is None:
            mismatches.append(
                Mismatch(
                    kind="MISSING_SHEET",
                    sheet=name,
                    message=f"文件里没有「{name}」这张表,页面预览有",
                )
            )
            continue

        # ---- 列:先比集合,再比顺序。两者的修法不同,所以分开报
        if set(want.headers) != set(got.headers):
            missing = [h for h in want.headers if h not in got.headers]
            extra = [h for h in got.headers if h not in want.headers]
            mismatches.append(
                Mismatch(
                    kind="HEADER_SET",
                    sheet=name,
                    expected=", ".join(want.headers),
                    actual=", ".join(got.headers),
                    message=(
                        f"文件缺列 {missing or '无'};多列 {extra or '无'}。"
                        "平台按列头取值,缺一列就是一个字段没提交"
                    ),
                )
            )
        elif want.headers != got.headers:
            mismatches.append(
                Mismatch(
                    kind="HEADER_ORDER",
                    sheet=name,
                    expected=", ".join(want.headers),
                    actual=", ".join(got.headers),
                    message="列集合相同但顺序不同。下面的单元格差异大概率都是它造成的",
                )
            )

        if len(want.rows) != len(got.rows):
            mismatches.append(
                Mismatch(
                    kind="ROW_COUNT",
                    sheet=name,
                    expected=str(len(want.rows)),
                    actual=str(len(got.rows)),
                    message=(
                        "行数不一致。每一格都对但少一行,等于少上架一个 SKU —— "
                        "这种缺陷逐格看是看不出来的"
                    ),
                )
            )

        # ---- 单元格:按**列头**定位,不按列序号
        #
        # 列序号定位在列错位时会把整行判成全错,而且错得很安静。按列头定位
        # 则只报真正变了值的那一格,列错位本身由上面的 HEADER_ORDER 单独说。
        want_index = {h: i for i, h in enumerate(want.headers)}
        got_index = {h: i for i, h in enumerate(got.headers)}
        shared = [h for h in want.headers if h in got_index]

        for row_no, want_row in enumerate(want.rows, start=1):
            if row_no > len(got.rows):
                break
            got_row = got.rows[row_no - 1]
            for header in shared:
                wi, gi = want_index[header], got_index[header]
                want_cell = want_row[wi] if wi < len(want_row) else ""
                got_cell = got_row[gi] if gi < len(got_row) else ""
                checked += 1
                if want_cell == got_cell:
                    continue
                cell_problems += 1
                if cell_problems > max_cell_mismatches:
                    continue
                mismatches.append(
                    Mismatch(
                        kind="CELL",
                        sheet=name,
                        row=row_no,
                        column=header,
                        expected=want_cell,
                        actual=got_cell,
                        message=_explain_cell(want_cell, got_cell),
                    )
                )

    if cell_problems > max_cell_mismatches:
        mismatches.append(
            Mismatch(
                kind="CELL",
                sheet="—",
                message=(
                    f"另有 {cell_problems - max_cell_mismatches} 处单元格差异未列出"
                    f"(上限 {max_cell_mismatches})。先修上面的列或行问题再重跑"
                ),
            )
        )

    return ParityReport(
        mismatches=tuple(mismatches),
        checked_cells=checked,
        sheets=tuple(expected),
    )


def _explain_cell(want: str, got: str) -> str:
    """给一格差异配一句诊断。

    不是所有差异都同等神秘。下面这几类有明确的根因,直接说出来比让人
    盯着两个字符串猜要快得多 —— 尤其是空格和公式这两种肉眼看不出的。
    """
    if want and not got:
        return "页面有值,文件是空的"
    if got and not want:
        return "文件有值,页面是空的"
    if want.strip() == got.strip():
        return "只差首尾空白。平台按字符数截断,空白也算长度"
    if got.startswith("'") and got[1:] == want:
        # 反过来的情况(页面带引号、文件不带)不单列:那说明预览侧被转义了,
        # 而预览是给人看的,不该有转义痕迹,走下面的通用分支即可
        return (
            "文件里多了一个前导单引号。CSV 的公式转义用的是这个,"
            "但 xlsx 不该这么写 —— Excel 会把引号本身显示出来"
        )
    if want.startswith(("=", "+", "-", "@")):
        return (
            "页面值以公式字符开头,而文件里的值不同。检查写入时是否被"
            "当成公式(openpyxl 会把 `=` 开头的字符串写成 formula)"
        )
    if len(want) != len(got):
        return f"长度不同:页面 {len(want)} 字符,文件 {len(got)} 字符"
    return "值不同"


# ==========================================================
# 读真实文件(需要 openpyxl,所以隔在最后)
# ==========================================================


def read_workbook(data: bytes) -> dict[str, SheetTable]:
    """把导出的 xlsx 字节读成若干张表。

    openpyxl 是延迟导入的,理由与 `export_writer.to_xlsx` 相同:
    上面那些判定函数会被纯测试大量导入,不该为一个只在这里用的依赖买单。

    `data_only=True`:如果某一格真的被写成了公式,这里读到的是缓存值
    (我们自己写的文件里没有缓存值,于是是 `None`),于是差异会以
    "页面有值,文件是空的"的形式暴露出来 —— 那正是需要被看见的故障。
    """
    import io

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 环境问题,不是逻辑问题
        from app.core.errors import ConfigError

        raise ConfigError(
            "比对 Excel 需要 openpyxl,当前环境没有安装。先 pip install openpyxl"
        ) from exc

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    tables: dict[str, SheetTable] = {}
    try:
        for sheet in workbook.worksheets:
            rows = [
                tuple(normalize(c) for c in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            headers = rows[0] if rows else ()
            body = tuple(r for r in rows[1:] if any(c != "" for c in r))
            tables[sheet.title] = SheetTable(
                name=sheet.title, headers=tuple(headers), rows=body
            )
    finally:
        workbook.close()
    return tables
